# -*- coding: utf-8 -*-

import os
import csv
import requests
import re
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
import openpyxl
import openpyxl.cell
import openpyxl.worksheet
import openpyxl.drawing
import openpyxl.drawing.image
import openpyxl.utils.datetime
import openpyxl.styles.numbers
import datetime
from .utils import convert_datetime_string_to_unix_local
from django.shortcuts import render
from django.http import HttpResponse

from openpyxl.chart import (
    PieChart,
    Reference
)


def csv_clean(input_string):
    new_string = str(input_string).strip()
    new_string = '"' + new_string + '"'
    if new_string.startswith('"http'):
        new_string = "=HYPERLINK(" + new_string + ")"

    return new_string

def filter(connections, user_id):
    clean_connections = set()
    for connection in connections:
        source_id, target_id = connection['source_id'], connection['target_id']
        if source_id == user_id:
            clean_connections.add(target_id)
        else:
            clean_connections.add(source_id)

    return clean_connections

def populate_sponsor_notifications(sponsors):
    sponsor_notifications = []
    temp_api = SocioRestApi()
    for sponsor in sponsors:
        notifications, valid = temp_api.get_sponsor_notifications(sponsor['id'])
        if valid:
            for notification in notifications:
                sponsor_notifications.append(notification)
        else:
            print("Failed to Get Sponsor Notifications")
            exit(1)

        return sponsor_notifications

def gather_notifications(event_notifications, sponsor_notifications):
    notifications = []

    if event_notifications is not None:
        for event_notif in event_notifications:
            new_dict = {}
            new_dict['content'] = event_notif['content']
            new_dict['post_time'] = event_notif['post_time']
            notifications.append(new_dict)

    if sponsor_notifications is not None:
        for sponsor_notification in sponsor_notifications:
            new_dict = {}
            new_dict['content'] = sponsor_notification['content']
            new_dict['post_time'] = sponsor_notification['post_time']
            notifications.append(new_dict)

    return notifications


class SocioRestApi:
    def __init__(self):
        self.base_url = "https://socio-secondary.herokuapp.com"

    def get_event(self, event_id):
        response = requests.get(self.base_url + "/events/" + str(event_id))
        return response.json(), response.status_code == requests.codes.ok

    def get_event_attendees(self, event_id):
        response = requests.get(self.base_url + "/users/event/" + str(event_id))
        return response.json(), response.status_code == requests.codes.ok

    def get_user_accounts(self, user_id):
        response = requests.get(self.base_url + "/accounts/user/" + str(user_id))
        return response.json(), response.status_code == requests.codes.ok

    def get_user_connections(self, user_id, from_unix, to_unix):
        response = requests.get(
            self.base_url + "/connections/user/" + str(user_id) + "/" + str(from_unix) + "/" + str(to_unix))
        return response.json(), response.status_code == requests.codes.ok

    def get_user_linkedin_info(self, user_id):
        response = requests.get(
            self.base_url + "/info/user/" + str(user_id) + "/Linkedin"
        )
        return response.json(), response.status_code == requests.codes.ok

    def get_event_location(self, event_id):
        response = requests.get(self.base_url + "/event_locations/" + str(event_id))
        return response.json(), response.status_code == requests.codes.ok

    def get_user(self, user_id):
        response = requests.get(self.base_url + "/users/" + str(user_id))
        return response.json(), response.status_code == requests.codes.ok

    def get_event_notifications(self, event_id):
        response = requests.get(self.base_url + "/event_notifications/event/" + str(event_id))
        return response.json(), response.status_code == requests.codes.ok

    def get_sponsor(self, event_id):
        response = requests.get(self.base_url + "/sponsors/event/" + str(event_id))
        return response.json(), response.status_code == requests.codes.ok

    def get_sponsor_notifications(self, sponsor_id):
        response = requests.get(self.base_url + "/sponsor_notifications/sponsor/" + str(sponsor_id))
        return response.json(), response.status_code == requests.codes.ok

class CsvBuilder:
    def __init__(self, output_directory, event_name, users_dict, sorting_key=lambda u : str(u['first_name'] + " " + u['last_name']).lower()):
        self.output_dir = output_directory
        self.event = event_name
        self.users_dict = users_dict
        self.sorting_key = sorting_key
        self.wb = Workbook()

    def output_to_csv(self, lines, identifier):
        output_file_name = str(self.output_dir) + "/" + str(self.event) + "_" + str(identifier) + ".csv"
        output_file_name = output_file_name.replace("/",".")
        with open(output_file_name,'wb') as f:
            f.write(lines.encode('UTF-8'))
            
        return output_file_name

    def create_linkedin_info(self, users, key=None):
        if key is None:
            key = self.sorting_key
        location_regex = re.compile('name=\"([\w\/,\s]+)\"')
        output_lines = str()

        output_lines += "Attendee Name,Last Name,Industry,Position,Company,Location,Number of LI connections,Headline" + "\n"

        for u in sorted(users, key=key):
            li_info = u['linkedin_info']
            company = li_info['company'] if 'company' in li_info and li_info['company'] is not None else 'N/A'
            headline = li_info['headline'] if 'headline' in li_info and li_info['headline'] is not None else 'N/A'
            industry = li_info['industry'] if 'industry' in li_info and li_info['industry'] is not None else 'N/A'
            location = li_info['location'] if 'location' in li_info and li_info['location'] is not None else 'N/A'
            num_connections = str(li_info['num_connections']) if 'num_connections' in li_info and li_info['num_connections'] is not None else 'N/A'
            if num_connections == "500":
                num_connections = "500+"
            position = li_info['position'] if 'position' in li_info and li_info['position'] is not None else 'N/A'
            if(location is not 'N/A'):
                location = location_regex.search(location).group(1)
            try:
                ordered = csv_clean(u['first_name']), csv_clean(u['last_name']), csv_clean(industry),csv_clean(position),csv_clean(company),csv_clean(location), csv_clean(num_connections), csv_clean(headline)
            except AttributeError:
                print("Error")
            output_lines += ",".join(ordered) + "\n"

        output_lines += "\n" + csv_clean("NOTE: What were the demographics of your attendees (i.e location, gender, interests, industry etc.)? " \
                        "For more info visit socioevents.com or reach us at platform@atsocio.com!\n")

        return self.output_to_csv(output_lines, 'linkedin_info')

    def create_complete_contact_list(self, users, key=None):
        if key is None:
            key = self.sorting_key

        output_lines = str()
        max_size = 0
        for u in sorted(users, key=key):
            line = csv_clean(u['first_name']) + "," + csv_clean(u['last_name'])
            if len(u['accounts']) > max_size:
                max_size = len(u['accounts'])
            for account in u['accounts']:
                line += "," + csv_clean(account['type']) + "," + csv_clean(account['new_detail'])

            output_lines += line + "\n"

        header_line = "Attendee Name,Last Name"
        for i in range(0, max_size):
            header_line += "," + "INFO " + str(i + 1) + ","

        output_lines = header_line + "\n" + output_lines
        output_lines += "\n" + csv_clean("NOTE: What are the contact info (i.e social media profiles, emails, phone "
                                           "numbers etc.) of all the attendees?"
                                           " For more info visit socioevents.com or reach us at platform@atsocio.com!\n")
        return self.output_to_csv(output_lines, 'complete_contact_list')

    def create_influencer_tree(self, users, key=None):
        if key is None:
            key = self.sorting_key

        output_lines = str()

        output_lines += "Con. Heat,Connector,Connected" + "\n"

        id_lookup = {}
        for u in users:
            id_lookup[u['id']] = u['first_name'] + " " + u['last_name']

        pitty_users = set()
        for u in sorted(users, key=key):
            if len(u['connections']) == 0:
                pitty_users.add(u['id'])
                continue

            output_lines += "," + csv_clean(id_lookup[u['id']])
            first_connection = True
            for connection in u['connections']:
                if not first_connection:
                    output_lines += ","
                else:
                    first_connection = False

                if connection not in id_lookup.keys():
                    temp_api = SocioRestApi()
                    c_user, valid = temp_api.get_user(connection)
                    if not valid:
                        print("Could not find the unattended user in database")
                        exit(1)
                    id_lookup[connection] = c_user['first_name'] + " " + c_user['last_name']

                output_lines += "," + csv_clean(id_lookup[connection]) + "\n"

        for user_id in pitty_users:
            output_lines += "," + csv_clean(id_lookup[user_id]) + ",\n"

        return self.output_to_csv(output_lines, 'influencer_tree')

    def create__notifications_overview(self, notifications, notif_count):
        output_lines = str()
        output_lines += "Notification Time:,Notification Message:\n"
        for notification in notifications:
            notification_date = datetime.datetime.fromtimestamp(notification['post_time'])
            output_lines += str(notification_date.strftime("%H:%M")) + "," + csv_clean(str(notification['content']))
            output_lines += "\n"

        output_lines += "\n\nNotifications received," + str(notif_count) + "\n"
        output_lines += "\n\nNOTE: Schedules and messages of notifications that were sent by the event planner."

        return self.output_to_csv(output_lines, 'notifications')

    def create_metrics_overview(self, total_connections, total_accounts_shared, event_location):
        output_lines = str()
        output_lines += "Total number of new connections:," + str(total_connections) + "\n"
        output_lines += "Total number of accounts shared:," + str(total_accounts_shared) + "\n"
        start_date = datetime.datetime.fromtimestamp(event_location['start_time'])
        end_date = datetime.datetime.fromtimestamp(event_location['end_time'])
        output_lines += "Start time," + str(start_date.strftime("%H:%M")) + "\n"
        output_lines += "End time," + str(end_date.strftime("%H:%M")) + "\n\n"

        output_lines += "\n"

        output_lines += csv_clean("NOTE: Who has connected with whom during the event? Who were the influencers? "
                                  "How many connections were made during the event? For more info visit "
                                  "socioevents.com or reach us at platform@atsocio.com!")

        return self.output_to_csv(output_lines, 'metrics')

    def create_platform_demographics(self, users):
        output_lines = str()
        accounts_counter = {}
        for u in users:
            for a in u['accounts']:
                if a['type'] not in accounts_counter.keys():
                    accounts_counter[a['type']] = 1
                else:
                    accounts_counter[a['type']] += 1

        for account_type, count in accounts_counter.items():
            output_lines += account_type + "," + str(count) + "\n"


        return self.output_to_csv(output_lines, 'platform_demographics')
        
    def create_easy_mailing_list(self, users, key=None):
        output_lines = str()
        if key is None:
            key = self.sorting_key
        max_size = 0
        for u in sorted(users, key=key):
            line = csv_clean(u['first_name']) + "," + csv_clean(u['last_name'])
            email_count = 0
            for a in u['accounts']:
                if a['type'] == 'Email':
                    email_count += 1
                    line += "," + a['new_detail']
            if email_count > max_size:
                max_size = email_count
            output_lines += line + "\n"

        header_line = "Attendee Name,Last Name"
        for i in range(0, max_size):
            header_line += "," + "Email " + str(i + 1)

        output_lines = header_line + "\n" + output_lines
        return self.output_to_csv(output_lines, 'easy_mailing_list')

class AccountsUriBuilder:
    def __init__(self):
        self.lookups = {
            'Facebook': {
                'insert_string': 'https://fb.com/{i}',
                'regex_string': r'^((https?:\/\/)?(www\.)?(facebook|fb)\.com\/)?([^\s\/]+)$',
                'regex_group': 5
            },
            'Twitter': {
                'insert_string': 'https://twitter.com/{i}',
                'regex_string': r'^((https?:\/\/)?(www\.)?(twitter)\.com\/)?@?([^\s\/]+)$',
                'regex_group': 5
            },
            'Instagram': {
                'insert_string': 'https://instagram.com/{i}',
                'regex_string': r'^((https?:\/\/)?(www\.)?(instagram)\.com\/)?([^\s\/]+)$',
                'regex_group': 5
            },
            'Snapchat': {
                'insert_string': '{i}',
                'regex_string': r'^([\w]+)$',
                'regex_group': 0
            },
            'Linkedin': {
                'insert_string': 'https://linkedin.com/in/{i}',
                'regex_string': r'^((https?:\/\/)?(www\.)?(linkedin)\.com\/in\/)([^\s\/]+)$',
                'regex_group': 5
            },
            'Email': {
                'insert_string': '{i}',
                'regex_string': r'^(.+\@.+\..+)$',
                'regex_group': 0
            },
            'Phone': {
                'converter': lambda x: x.replace("(", "").replace(")", "")
                    .replace(" ", "").replace("-", "").replace("/", "").replace(".", ""),
                'insert_string': '{i}',
                'regex_string': r'^([\d#]+)$',
                'regex_group': 0
            },
            'Skype': {
                'insert_string': '{i}',
                'regex_string': r'^([^\s]+)$',
                'regex_group': 0
            },
            'Swarm': {
                'insert_string': '{i}',
                'regex_string': r'^([^\s]+)$',
                'regex_group': 0
            },
            'Pinterest': {
                'insert_string': 'https://pinterest.com/{i}',
                'regex_string': r'^((https?:\/\/)?(www\.)?(pinterest)\.com\/)?([^\s\/]+)$',
                'regex_group': 5
            },
            'Vine': {
                'insert_string': '{i}',
                'regex_string': r'^([^\s]+)$',
                'regex_group': 0
            },
            'Periscope': {
                'insert_string': '{i}',
                'regex_string': r'^([^\s]+)$',
                'regex_group': 0
            },
            'Google+': {
                'insert_string': 'https://plus.google.com/{i}',
                'regex_string': r'^((https?:\/\/)?(www\.)?(plus.google)\.com\/)?([^\s\/]+)$',
                'regex_group': 5
            },
            'Website': {
                'insert_string': 'http://www.{i}',
                'regex_string': r'^(https?:\/\/)?((www\.)?[^\s]+\.[^\s]+)$',
                'regex_group': 2
            },
            'LinkedinCompany': {
                'insert_string': 'https://linkedin.com/company/{i}',
                'regex_string': r'^((https?:\/\/)?(www\.)?(linkedin)\.com\/company\/)([^\s\/]+)$',
                'regex_group': 5
            },
            'Youtube': {
                'insert_string': 'https://youtube.com/user/{i}',
                'regex_string': r'^((https?:\/\/)?(www\.)?(youtube)\.com\/user\/)([^\s\/]+)$',
                'regex_group': 5
            }
        }

    def convert(self, type, identifier):
        type_strip = type.strip()
        identifier_strip = identifier.strip()
        if type_strip not in self.lookups:
            raise ValueError("Invalid Type: " + type_strip)

        lookup = self.lookups.get(type_strip)

        if 'converter' in lookup:
            converter = lookup.get('converter')
            identifier_strip = converter(identifier_strip)

        new_detail = lookup.get('insert_string').replace('{i}', identifier_strip)

        if not re.compile(lookup.get('regex_string')).match(new_detail):
            raise ValueError("Created an Invalid Detail '" + new_detail + "' from '" + identifier_strip + "'")

        return new_detail

class ExcelBuilder:
    AttendeeDemographics = "Attendee Demographics"
    CompleteContactsList = "Complete Contacts List"
    InfluencerTree = "Cons & Infl. Tree and Metrics"
    PlatformDemographics = "Platform Demographics"
    EasyMailingList = "Easy Mailing List"

    def __init__(self):
        self.wb = Workbook()
        self.worksheet_definitions = {
            ExcelBuilder.AttendeeDemographics: {
                "sheetNumber": 0,
                "sheetTitle": "Attendee Demographics",
                "columnsToSkip": set()
            },
            ExcelBuilder.CompleteContactsList: {
                "sheetNumber": 1,
                "sheetTitle": "Complete Contacts List",
                "columnsToSkip": set()
            },
            ExcelBuilder.InfluencerTree: {
                "sheetNumber": 2,
                "sheetTitle": "Cons & Infl. Tree and Metrics",
                "columnsToSkip": set()
            },
            ExcelBuilder.PlatformDemographics: {
                "sheetNumber": 3,
                "sheetTitle": "Platform Demographics",
                "columnsToSkip": set()
            },
            ExcelBuilder.EasyMailingList: {
                "sheetNumber": 4,
                "sheetTitle": "Easy Mailing List",
                "columnsToSkip": set()
            }
        }

        self.worksheets = {}
        for worksheet, ws_definition in sorted(self.worksheet_definitions.items(),
                                               key=lambda dict_item: dict_item[1]['sheetNumber']):
            if ws_definition['sheetNumber'] == 0:
                this_worksheet = self.wb.active
                this_worksheet.title = ws_definition['sheetTitle']
            else:
                this_worksheet = self.wb.create_sheet(
                    title=ws_definition['sheetTitle'],
                    index=ws_definition['sheetNumber']
                )

            self.worksheets[worksheet] = this_worksheet

    def set_cell_formatting(self, cell, font=None, fill=None, border=None, alignment=None, number_format=None):
        if not isinstance(cell, openpyxl.cell.Cell):
            print("Not a Cell!: " + str(cell))
            return

        if font is not None:
            cell.font = font

        cell.font = cell.font.copy(name='Calibri (Body)', sz=12)

        if fill is not None:
            cell.fill = fill

        if border is not None:
            cell.border = border

        if alignment is not None:
            cell.alignment = alignment

        if number_format is not None:
            cell.number_format = number_format

        if str(cell.value).isdigit():
            cell.value = int(cell.value)

    def set_column_width_to_max(self, worksheet, columns_to_skip=set()):

        if not isinstance(worksheet, openpyxl.worksheet.Worksheet):
            print("Error: not a worksheet: " + str(worksheet))
            return

        dimensions = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value is not None and "NOTE:" not in str(cell.value):
                    try:
                        this_cell_width = len(cell.value)
                        if str(cell.value).startswith("=HYPERLINK("):
                            this_cell_width -= 13

                        if this_cell_width < 0:
                            this_cell_width = 0

                        dimensions[cell.column] = max((dimensions.get(cell.column, 0), this_cell_width))
                    except TypeError:
                        pass

        for column, value in dimensions.items():
            if column in columns_to_skip:
                continue
            worksheet.column_dimensions[column].width = value + 5

    def create_attendee_demographics_from_csv(self, csv_file_location):
        with open(csv_file_location) as f:
            csv_reader = csv.reader(f)
            ws = self.worksheets[ExcelBuilder.AttendeeDemographics]

            for row_index, row in enumerate(csv_reader):
                for column_index, csv_cell in enumerate(row):
                    column_letter = get_column_letter(column_index + 1)
                    row_lookup_index = str(row_index + 1)
                    this_cell = ws[str(column_letter + row_lookup_index)]

                    if not isinstance(this_cell, openpyxl.cell.Cell):
                        print("what the hell is going on. not a cell?!: " + str(this_cell))
                        continue

                    this_cell.value = csv_cell

                    font = None
                    fill = None
                    border = None
                    alignment = None
                    number_format = None

                    # Header
                    if row_index == 0:
                        font = Font(
                            bold=True,
                            color=colors.RED
                        )
                        grey = "D9D9D9"
                        fill = PatternFill(
                            fill_type='solid',
                            start_color=grey,
                            end_color=grey
                        )
                    # First and Last Names
                    elif row_index > 0 and column_index < 2:
                        font = Font(
                            bold=True
                        )
                    # Number of Linkedin Connections
                    elif row_index > 0 and column_index == 6:
                        alignment = Alignment(horizontal='center')

                    if "NOTE:" in csv_cell:
                        alignment = Alignment(
                            horizontal='left',
                            wrap_text=True,
                            vertical='top'
                        )

                    self.set_cell_formatting(this_cell, font, fill, border, alignment, number_format)

            last_row = int(row_lookup_index)
            ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row + 2, end_column=3)

        os.remove(csv_file_location)

    def create_complete_contacts_list_from_csv(self, csv_file_location):
        with open(csv_file_location) as f:
            csv_reader = csv.reader(f)
            ws = self.worksheets[ExcelBuilder.CompleteContactsList]

            info_type_cell_columns = set()
            number_of_rows = 0
            for row_index, row in enumerate(csv_reader):
                number_of_rows += 1
                for column_index, csv_cell in enumerate(row):
                    column_letter = get_column_letter(column_index + 1)
                    row_lookup_index = str(row_index + 1)
                    this_cell = ws[str(column_letter + row_lookup_index)]

                    if not isinstance(this_cell, openpyxl.cell.Cell):
                        print("what the hell is going on. not a cell?!: " + str(this_cell))
                        continue

                    this_cell.value = csv_cell

                    font = None
                    fill = None
                    border = None
                    alignment = None
                    number_format = None

                    #Header
                    if row_index == 0:
                        font = Font(
                            bold=True,
                            color=colors.RED
                        )
                        grey = "D9D9D9"
                        fill = PatternFill(
                            fill_type='solid',
                            start_color=grey,
                            end_color=grey
                        )
                    # First and Last Names
                    elif row_index > 0 and column_index < 2:
                        font = Font(
                            bold=True
                        )
                    # Info Columns
                    elif row_index > 0 and column_index > 1 and column_index % 2 == 0:
                        info_type_cell_columns.add(column_letter)
                    # Handle Columns
                    else:
                        alignment = Alignment(
                            horizontal='left'
                        )
                        if "HYPERLINK" in csv_cell:
                            font = Font(
                                color=colors.BLUE
                            )

                    if "NOTE:" in csv_cell:
                        alignment = Alignment(
                            horizontal='left',
                            wrap_text=True,
                            vertical='top'
                        )

                    self.set_cell_formatting(this_cell, font, fill, border, alignment, number_format)

            last_row = int(row_lookup_index)
            ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row + 2, end_column=4)

            for column in info_type_cell_columns:
                for row_index in range(1, number_of_rows - 2):
                    info_cell = ws[str(column) + str(row_index + 1)]
                    fill = PatternFill(fill_type='solid', start_color=colors.YELLOW, end_color=colors.YELLOW)
                    self.set_cell_formatting(info_cell, fill=fill)

        os.remove(csv_file_location)

    def create_influencer_tree_from_csv(self, csv_file_location, metrics_file, notifications_file):
        with open(csv_file_location) as f:
            csv_reader = csv.reader(f)
            ws = self.worksheets[ExcelBuilder.InfluencerTree]
            if not isinstance(ws, openpyxl.worksheet.Worksheet):
                print("not a worksheet: " + str(ws))
                return
            number_of_rows = 0
            start_of_connector = []

            for row_index, row in enumerate(csv_reader):
                number_of_rows += 1
                for column_index, csv_cell in enumerate(row):
                    column_letter = get_column_letter(column_index + 1)
                    row_lookup_index = str(row_index + 1)
                    this_cell = ws[str(column_letter + row_lookup_index)]

                    if not isinstance(this_cell, openpyxl.cell.Cell):
                        print("what the hell is going on. not a cell?!: " + str(this_cell))
                        continue

                    this_cell.value = csv_cell

                    font = None
                    fill = None
                    border = None
                    alignment = None
                    number_format = None

                    #Header
                    if row_index == 0:
                        font = Font(
                            bold=True,
                            color=colors.RED
                        )
                        grey = "D9D9D9"
                        fill = PatternFill(
                            fill_type='solid',
                            start_color=grey,
                            end_color=grey
                        )
                    # Connectors
                    if row_index > 0 and column_index == 1 and this_cell.value is not None and len(str(this_cell.value).strip()) > 0:
                        start_of_connector.append(int(row_lookup_index))
                        font = Font(
                            bold=True
                        )
                        alignment = Alignment(
                            horizontal='center',
                            vertical='center'
                        )

                    self.set_cell_formatting(this_cell, font, fill, border, alignment, number_format)
            # Merge cells for each 'connector'
            start_of_connector.append(number_of_rows + 1)
            for i in range(0, len(start_of_connector) - 1):
                start = start_of_connector[i]
                end = start_of_connector[i + 1] - 1
                ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)

            # Add Metrics Data
            start_column = 5
            start_row = 4
            even = 1
            with open(metrics_file) as m:
                metrics_reader = csv.reader(m)
                for row_index, row in enumerate(metrics_reader):
                    for column_index, csv_cell in enumerate(row):
                        alignment = None

                        if str(csv_cell).startswith("NOTE:"):
                            start_row += 1
                            alignment = Alignment(
                                horizontal='left',
                                wrap_text=True,
                                vertical='top'
                            )

                        column = get_column_letter(start_column)
                        ws[str(column + str(start_row))] = csv_cell
                        this_cell = ws[str(column + str(start_row))]

                        font = None
                        fill = None

                        if even == 0:
                            even = 1
                            start_column -= 1
                            start_row += 1

                            alignment = Alignment(
                                horizontal='right'
                            )
                        else:
                            even = 0
                            start_column += 1

                            font = Font(
                                bold=True,
                                color=colors.BLACK
                            )
                            grey = "D9D9D9"
                            fill = PatternFill(
                                fill_type='solid',
                                start_color=grey,
                                end_color=grey
                            )
                            if str(csv_cell).startswith("NOTE:"):
                                fill = None

                        self.set_cell_formatting(this_cell, font, fill, border, alignment, number_format)

                ws.merge_cells(start_row=int(start_row), start_column=5, end_row=int(start_row) + 2, end_column=9)

            # Add Notifications Data
            start_column = 5
            start_row = 16
            even = 1
            with open(notifications_file) as n:
                notifications_reader = csv.reader(n)
                for row_index, row in enumerate(notifications_reader):
                    for column_index, csv_cell in enumerate(row):

                        if str(csv_cell).startswith("NOTE:"):
                            start_row += 1

                        if str(csv_cell).startswith("Notifications received"):
                            start_row += 1

                        column = get_column_letter(start_column)
                        ws[str(column + str(start_row))] = csv_cell
                        this_cell = ws[str(column + str(start_row))]

                        alignment = None
                        font = None
                        fill = None

                        if even == 0:
                            even = 1
                            start_column -= 1
                            start_row += 1
                        else:
                            even = 0
                            start_column += 1

                            grey = "D9D9D9"
                            fill = PatternFill(
                                fill_type='solid',
                                start_color=grey,
                                end_color=grey
                            )
                            alignment = Alignment(
                                horizontal='center'
                            )

                        if start_row == 16:
                            font = Font(
                                bold=True,
                                color=colors.BLACK
                            )
                            fill = None

                        if str(csv_cell).startswith("Notifications received") \
                                or str(csv_cell).startswith("Notification Message"):
                            alignment = None
                            font = Font(
                                bold=True,
                                color=colors.BLACK
                            )

                        if str(csv_cell).startswith("NOTE:"):
                            fill = None
                            alignment = Alignment(
                                horizontal='left',
                                wrap_text=True,
                                vertical='top'
                            )
                            font = Font(
                                bold=True,
                                color=colors.BLACK
                            )

                        self.set_cell_formatting(this_cell, font, fill, border, alignment, number_format)

                ws.merge_cells(start_row=int(start_row), start_column=5, end_row=int(start_row) + 2, end_column=9)

            self.worksheet_definitions.get(ExcelBuilder.InfluencerTree).get('columnsToSkip').add('F')

            img = openpyxl.drawing.image.Image('./res/influencer_heat.png', nochangeaspect=False)
            img.drawing.height = 21.3378 * (end - 1)
            img.drawing.width = 112
            ws.add_image(img, 'A2')

        os.remove(csv_file_location)
        os.remove(metrics_file)
        os.remove(notifications_file)

    def create_platform_demographics_from_csv(self, csv_file_location):
        with open(csv_file_location) as f:
            csv_reader = csv.reader(f)
            ws = self.worksheets[ExcelBuilder.PlatformDemographics]
            if not isinstance(ws, openpyxl.worksheet.Worksheet):
                print("not a worksheet: " + str(ws))
                return
            number_of_rows = 0

            ws['A1'] = str(ExcelBuilder.PlatformDemographics)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
            font = Font(
                bold=True,
                color=colors.RED
            )
            grey = "D9D9D9"
            fill = PatternFill(
                fill_type='solid',
                start_color=grey,
                end_color=grey
            )
            alignment = Alignment(
                horizontal='center',
                vertical='center'
            )
            self.set_cell_formatting(ws['A1'], font, fill, None, alignment, None)

            for row_index, row in enumerate(csv_reader):
                number_of_rows += 1
                for column_index, csv_cell in enumerate(row):
                    column_letter = get_column_letter(column_index + 1)
                    row_lookup_index = str(row_index + 2)
                    this_cell = ws[str(column_letter + row_lookup_index)]

                    if not isinstance(this_cell, openpyxl.cell.Cell):
                        print("what the hell is going on. not a cell?!: " + str(this_cell))
                        continue

                    this_cell.value = csv_cell

                    font = None
                    fill = None
                    border = None
                    alignment = None
                    number_format = None

                    if column_index == 1 and str(csv_cell).isdigit():
                        number_format = openpyxl.styles.numbers.FORMAT_NUMBER
                        this_cell.value = int(csv_cell)

                    self.set_cell_formatting(this_cell, font, fill, border, alignment, number_format)

            # Create and add pie chart
            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=number_of_rows + 1)
            data = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=number_of_rows + 1)
            pie.add_data(data)
            pie.set_categories(labels)
            pie.title = None
            pie.style = 2
            pie.height = 13
            pie.width = 21.55

            ws.add_chart(pie, anchor='A2')

            # Add the last note

            ws['A28'] = "NOTE: Which platforms are being utilized the most by the attendees? For more info visit " \
                        "socioevents.com or reach us at platform@atsocio.com!"

            ws.merge_cells(start_row=28, start_column=1, end_row=30, end_column=8)
            font = Font(
                bold=True,
            )
            alignment = Alignment(
                horizontal='left',
                wrap_text=True,
                vertical='top'
            )
            self.set_cell_formatting(ws['A28'], font, fill, None, alignment, None)

        os.remove(csv_file_location)

    def create_easy_mailing_list_from_csv(self, csv_file_location):
        with open(csv_file_location) as f:
            csv_reader = csv.reader(f)
            ws = self.worksheets[ExcelBuilder.EasyMailingList]

            for row_index, row in enumerate(csv_reader):
                for column_index, csv_cell in enumerate(row):
                    column_letter = get_column_letter(column_index + 1)
                    row_lookup_index = str(row_index + 1)
                    this_cell = ws[str(column_letter + row_lookup_index)]

                    if not isinstance(this_cell, openpyxl.cell.Cell):
                        print("what the hell is going on. not a cell?!: " + str(this_cell))
                        continue

                    this_cell.value = csv_cell

                    font = None
                    fill = None
                    border = None
                    alignment = None
                    number_format = None

                    #Header
                    if row_index == 0:
                        font = Font(
                            bold=True,
                            color=colors.RED
                        )
                        grey = "D9D9D9"
                        fill = PatternFill(
                            fill_type='solid',
                            start_color=grey,
                            end_color=grey
                        )
                    # First and Last Names
                    elif row_index > 0 and column_index < 2:
                        font = Font(
                            bold=True
                        )

                    self.set_cell_formatting(this_cell, font, fill, border, alignment, number_format)

        os.remove(csv_file_location)

    def save(self, output_directory, output_file):
        for title, ws in self.worksheets.items():
            if not isinstance(ws, openpyxl.worksheet.Worksheet):
                print("Error: Not a workbook: " + str(ws))
                continue

            columns_to_skip = set()
            if title in self.worksheet_definitions.keys():
                columns_to_skip = self.worksheet_definitions[title].get('columnsToSkip')

            self.set_column_width_to_max(ws, columns_to_skip)

        self.wb.save(output_directory + "/" + output_file + ".xlsx")

def perform_data_converstion(request):
    print("Started")
    event_id = 827
    total_connections_doubled = 0
    total_connections_actual = 0
    total_accounts = 0

    uriBuilder = AccountsUriBuilder()

    # Retrieve and Build Data Needed
    print("Retrieving Data for " + str(event_id))
    api = SocioRestApi()
    event, valid = api.get_event(event_id)
    if not valid:
        print("Failed To Get Event")
        exit(1)
    attendees, valid = api.get_event_attendees(event_id)
    if not valid:
        print("Failed To Get Event Attendees")
        exit(1)
    event_location, valid = api.get_event_location(event_id)
    if not valid:
        print("Failed To Get Event Location")
        exit(1)

    event_notifications, valid = api.get_event_notifications(event_id)
    if not valid:
        print("Failed to Get Event Notifications")
        exit(1)

    sponsors, valid = api.get_sponsor(event_id)
    if valid:
        sponsor_notifications = populate_sponsor_notifications(sponsors)
    else:
        print("Failed to Get Sponsor Notifications")
        exit(1)

    notifications = gather_notifications(event_notifications, sponsor_notifications)

    user_dict = {}
    for user in attendees:
        if user['id'] in user_dict.keys():
            raise KeyError("Attendee already exists")
        user_dict[user['id']] = user

        linkedin_info, valid = api.get_user_linkedin_info(user['id'])
        if not valid:
            print('Failed to retrieve linkedin info for ' + str(user, 'utf-8'))
            exit(1)
        user['linkedin_info'] = linkedin_info[0] if len(linkedin_info) == 1 else {}

        connections, valid = api.get_user_connections(user['id'], event_location['start_time'] - 1800, event_location['end_time'])
        if not valid:
            print('Failed to retrieve connections for ' + str(user, 'utf-8'))
            exit(1)

        filtered_connections = filter(connections, user['id'])
        user['connections'] = filtered_connections
        accounts, valid = api.get_user_accounts(user['id'])

        if not valid:
            print('Failed to retrieve accounts for ' + str(user, 'utf-8'))
            exit(1)

        for account in accounts:
            account['new_detail'] = uriBuilder.convert(account['type'], account['detail'])
        user['accounts'] = accounts

        # Metrics Calculation
        for conn in filtered_connections:
            if conn not in user_dict:
                total_connections_actual += 1

        total_connections_doubled += len(filtered_connections)
        total_accounts += len(filtered_connections) * len(accounts)

    print("Creating CSVs")
    # Output to CSVs
    csvBuilder = CsvBuilder("./tmp", event["event_name"], user_dict)
    linkedin_info_file = csvBuilder.create_linkedin_info(attendees)
    complete_contact_list_file = csvBuilder.create_complete_contact_list(attendees)
    influencer_sorter = lambda u: float(1 / len(u['connections'])) if len(u['connections']) != 0 else float(0.0)
    influencer_tree_file = csvBuilder.create_influencer_tree(attendees, influencer_sorter)
    metrics_file = csvBuilder.create_metrics_overview(total_connections_actual, total_accounts, event_location)
    notifications_file = csvBuilder.create__notifications_overview(notifications, len(attendees) * len(notifications))
    platform_demographics_file = csvBuilder.create_platform_demographics(attendees)
    easy_mailing_list_file = csvBuilder.create_easy_mailing_list(attendees)

    print("Creating Excel File")
    # Merge CSVs into an Excel file
    excelBuilder = ExcelBuilder()
    excelBuilder.create_attendee_demographics_from_csv(linkedin_info_file)
    excelBuilder.create_complete_contacts_list_from_csv(complete_contact_list_file)
    excelBuilder.create_influencer_tree_from_csv(influencer_tree_file, metrics_file, notifications_file)
    excelBuilder.create_platform_demographics_from_csv(platform_demographics_file)
    excelBuilder.create_easy_mailing_list_from_csv(easy_mailing_list_file)
    excelBuilder.save("./tmp", event["event_name"])

    print("Completed")
    return HttpResponse('Completed : ' + str(event_id))




if __name__ == "__main__":
    # Get User Input
    event_id = int(input("Please enter an event id: "))
    perform_data_converstion(event_id=event_id)
